from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from core.models import ValidationResult

HEADERS = ["FECHA", "Codigo_Serie", "LECTOR", "ESTADO", "TOTAL_NUEVOS", "TOTAL_DUPLICADOS", "TOTAL_ERRORES"]


@dataclass(frozen=True)
class ExcelCounters:
    nuevos: int
    duplicados: int
    errores: int


class ExcelRepository:
    def __init__(self, path: Path, logger: logging.Logger | None = None) -> None:
        self.path = path
        self.logger = logger or logging.getLogger(__name__)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.ensure_exists()

    def ensure_exists(self) -> None:
        if self.path.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"
        ws.append(HEADERS)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        widths = {"A": 22, "B": 34, "C": 18, "D": 22, "E": 18, "F": 20, "G": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        wb.save(self.path)
        self.logger.info("Excel creado: %s", self.path)

    def get_new_codes(self) -> set[str]:
        self.ensure_exists()
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            code_idx = headers.index("Codigo_Serie")
            status_idx = headers.index("ESTADO")
            result: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[code_idx]
                status = row[status_idx]
                if code is not None and str(status).strip().upper() == "NUEVO":
                    result.add(str(code))
            return result
        finally:
            wb.close()

    def get_counters(self) -> ExcelCounters:
        self.ensure_exists()
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            status_idx = headers.index("ESTADO")
            nuevos = duplicados = errores = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                status = str(row[status_idx] or "").strip().upper()
                if status == "NUEVO":
                    nuevos += 1
                elif status == "DUPLICADO":
                    duplicados += 1
                elif status.startswith("ERROR"):
                    errores += 1
            return ExcelCounters(nuevos, duplicados, errores)
        finally:
            wb.close()

    def append_cycle(self, right: ValidationResult, left: ValidationResult, timestamp: datetime) -> None:
        self.ensure_exists()
        wb = load_workbook(self.path)
        try:
            ws = wb.active
            counters = self.get_counters()
            nuevos, duplicados, errores = counters.nuevos, counters.duplicados, counters.errores
            for result in (right, left):
                status = result.status.value
                if status == "NUEVO":
                    nuevos += 1
                elif status == "DUPLICADO":
                    duplicados += 1
                elif status.startswith("ERROR"):
                    errores += 1

                lector = "Derecho" if result.side.value == "right" else "Izquierdo"
                ws.append([
                    timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    result.code,
                    lector,
                    status,
                    nuevos,
                    duplicados,
                    errores,
                ])
                self._style_last_row(ws, status, result.code)
            wb.save(self.path)
            self.logger.info("Ciclo registrado en Excel: %s", self.path)
        except PermissionError:
            self.logger.exception("No se pudo escribir Excel. Probablemente está abierto: %s", self.path)
            raise
        finally:
            wb.close()

    @staticmethod
    def _style_last_row(ws, status: str, code: str) -> None:
        width = max(ws.column_dimensions["B"].width or 34, len(str(code)) + 4)
        ws.column_dimensions["B"].width = width

        fill = None
        if status == "DUPLICADO":
            fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        elif status.startswith("ERROR"):
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif status == "NUEVO":
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        if fill is not None:
            font = Font(color="000000", bold=status != "NUEVO")
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.font = font
