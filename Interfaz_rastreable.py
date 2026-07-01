import pandas as pd
import serial
import time
import os
import socket
import threading
import tkinter as tk
from tkinter import font
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ==========================================
# 1. CONFIGURACIÓN
# ==========================================
IP_SCANNER_DERECHO = '192.168.0.10'
IP_SCANNER_IZQUIERDO = '192.168.0.11'
PUERTO_SCANNERS = 5000  

PUERTO_PLC = 'COM3'    
ARCHIVO_EXCEL = 'base_datos_produccion.xlsx'
TIEMPO_ACTIVACION = 2  

# ==========================================
# 2. ESTADO COMPARTIDO (LÓGICA <-> INTERFAZ)
# ==========================================
estado_gui = {
    "plc_ok": False,
    "izq_ok": False,
    "der_ok": False,
    "izq_codigo": "ESPERANDO PIEZA...",
    "der_codigo": "ESPERANDO PIEZA...",
    "izq_color": "#444444", # Color foco apagado (Gris asfalto)
    "der_color": "#444444",
    "izq_verde_ts": 0,   # Timestamp para apagar el verde en 3s
    "der_verde_ts": 0,
    "kpi_nuevos": 0,
    "kpi_dup": 0,
    "kpi_err": 0,
    "mensaje_alerta": "" # Si no está vacío, muestra la pantalla gigante de error
}

# ==========================================
# 3. FUNCIONES DE EXCEL Y COMUNICACIÓN
# ==========================================
def verificar_y_crear_excel():
    if not os.path.exists(ARCHIVO_EXCEL):
        print(f"[INFO] Creando base de datos nueva con formato presentable...")
        wb = Workbook()
        ws = wb.active
        cabeceras = ['FECHA', 'Codigo_Serie', 'LECTOR', 'ESTADO', 'TOTAL_NUEVOS', 'TOTAL_DUPLICADOS', 'TOTAL_ERRORES']
        ws.append(cabeceras)
        
        fondo_cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        letra_cabecera = Font(color="FFFFFF", bold=True)
        for celda in ws[1]:
            celda.fill = fondo_cabecera
            celda.font = letra_cabecera
            
        anchos = {'A': 22, 'B': 30, 'C': 15, 'D': 22, 'E': 18, 'F': 20, 'G': 18}
        for col_letra, ancho in anchos.items():
            ws.column_dimensions[col_letra].width = ancho

        wb.save(ARCHIVO_EXCEL)
        estado_gui["kpi_nuevos"] = 0
        estado_gui["kpi_dup"] = 0
        estado_gui["kpi_err"] = 0
    else:
        try:
            df = pd.read_excel(ARCHIVO_EXCEL)
            if 'ESTADO' in df.columns:
                estado_gui["kpi_nuevos"] = len(df[df['ESTADO'] == 'NUEVO'])
                estado_gui["kpi_dup"] = len(df[df['ESTADO'] == 'DUPLICADO'])
                estado_gui["kpi_err"] = len(df[df['ESTADO'] == 'ERROR DE LECTURA'])
        except Exception:
            pass 

def registrar_en_excel(ws, codigo, lector, estado_lectura, fecha_hora):
    if estado_lectura == "NUEVO": estado_gui["kpi_nuevos"] += 1
    elif estado_lectura == "DUPLICADO": estado_gui["kpi_dup"] += 1
    elif estado_lectura == "ERROR DE LECTURA": estado_gui["kpi_err"] += 1

    ws.append([fecha_hora, codigo, lector, estado_lectura, estado_gui["kpi_nuevos"], estado_gui["kpi_dup"], estado_gui["kpi_err"]])
    
    if estado_lectura in ["DUPLICADO", "ERROR DE LECTURA"]:
        relleno_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        fuente_negra = Font(color="000000", bold=True)
        for celda in ws[ws.max_row]:
            celda.fill = relleno_rojo
            celda.font = fuente_negra

def armar_trama(valor_hex_little_endian):
    cmd = f"1100002{valor_hex_little_endian}"
    suma = sum(ord(c) for c in cmd) + 3
    checksum = f"{suma & 0xFF:02X}"
    return b'\x02' + cmd.encode() + b'\x03' + checksum.encode()

def enviar_dato_d0(plc, valor_hex_little_endian, descripcion):
    try:
        trama = armar_trama(valor_hex_little_endian)
        plc.write(b'\x05') 
        time.sleep(0.1)
        plc.write(trama)   
        time.sleep(0.1)
        plc.read(10)
        print(f"   [PLC] -> {descripcion} ejecutado con éxito.")
        return True
    except serial.SerialException:
        return False

# ==========================================
# 4. FUNCIONES DE AUTO-RECUPERACIÓN
# ==========================================
def reconectar_plc():
    try:
        p = serial.Serial(port=PUERTO_PLC, baudrate=9600, bytesize=serial.SEVENBITS, parity=serial.PARITY_EVEN, stopbits=serial.STOPBITS_ONE, timeout=1)
        return p
    except: return None

def reconectar_escaner(ip, puerto):
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.2)
        s.connect((ip, puerto))
        return s
    except: return None

# ==========================================
# 5. HILO DE PRODUCCIÓN (EL CEREBRO INVISIBLE)
# ==========================================
def hilo_produccion():
    print("\n--- INICIANDO LÓGICA DE PRODUCCIÓN ---")
    verificar_y_crear_excel()

    plc = None
    lector_derecho = None
    lector_izquierdo = None
    codigo_der = None
    codigo_izq = None

    while True:
        # --- MONITOREO DE CONEXIONES ---
        if plc is None:
            plc = reconectar_plc()
            if plc: enviar_dato_d0(plc, "0000", "Reset Inicial (D0 = 0)")
        
        if lector_derecho is None:
            lector_derecho = reconectar_escaner(IP_SCANNER_DERECHO, PUERTO_SCANNERS)
            
        if lector_izquierdo is None:
            lector_izquierdo = reconectar_escaner(IP_SCANNER_IZQUIERDO, PUERTO_SCANNERS)

        # Actualizar estado de la GUI
        estado_gui["plc_ok"] = (plc is not None)
        estado_gui["der_ok"] = (lector_derecho is not None)
        estado_gui["izq_ok"] = (lector_izquierdo is not None)

        # Generar alerta gigante si algo falta
        fallas = []
        if not estado_gui["plc_ok"]: fallas.append(f"PLC ({PUERTO_PLC})")
        if not estado_gui["der_ok"]: fallas.append(f"Escáner Derecho ({IP_SCANNER_DERECHO})")
        if not estado_gui["izq_ok"]: fallas.append(f"Escáner Izq ({IP_SCANNER_IZQUIERDO})")

        if fallas:
            estado_gui["mensaje_alerta"] = "¡MÁQUINA DETENIDA!\nEQUIPO DESCONECTADO:\n\n" + "\n".join(fallas)
            time.sleep(1)
            continue # Bloquea la lectura hasta que todo esté conectado
        else:
            estado_gui["mensaje_alerta"] = ""

        # --- LECTURA DE CÓDIGOS ---
        if not codigo_der:
            try:
                datos_der = lector_derecho.recv(1024)
                if datos_der: codigo_der = datos_der.decode('utf-8').strip()
                else: lector_derecho = None
            except socket.timeout: pass
            except Exception: lector_derecho = None 

        if not codigo_izq:
            try:
                datos_izq = lector_izquierdo.recv(1024)
                if datos_izq: codigo_izq = datos_izq.decode('utf-8').strip()
                else: lector_izquierdo = None
            except socket.timeout: pass
            except Exception: lector_izquierdo = None 

        # --- LÓGICA DE EVALUACIÓN ---
        if codigo_der and codigo_izq and plc and lector_derecho and lector_izquierdo:
            fecha_hora_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # FILTRO: Corrige etiquetas pegadas y limpia
            def limpiar_codigo(cod):
                c = cod.replace('/r', '\r').split('\r')[0]
                return c.strip()[:22]

            cod_der_limpio = limpiar_codigo(codigo_der)
            cod_izq_limpio = limpiar_codigo(codigo_izq)

            verificar_y_crear_excel()
            
            try:
                df = pd.read_excel(ARCHIVO_EXCEL)
                historial = df['Codigo_Serie'].astype(str).values
                
                def evaluar_codigo(cod):
                    if cod.upper() in ["ERROR", "NG", "NOREAD"] or not cod or len(cod) < 5:
                        return "ERROR DE LECTURA", "ERROR"
                    elif cod in historial:
                        return cod, "DUPLICADO"
                    else:
                        return cod, "NUEVO"

                resultado_der, estado_der = evaluar_codigo(cod_der_limpio)
                resultado_izq, estado_izq = evaluar_codigo(cod_izq_limpio)
                
                # Actualizar Textos en Pantalla
                estado_gui["der_codigo"] = f"{resultado_der}\n({estado_der})"
                estado_gui["izq_codigo"] = f"{resultado_izq}\n({estado_izq})"

                # Controlar Focos Izquierdos
                if estado_izq == "NUEVO":
                    estado_gui["izq_color"] = "#2ECC71" # Verde Neón
                    estado_gui["izq_verde_ts"] = time.time()
                else:
                    estado_gui["izq_color"] = "#E74C3C" # Rojo Neón

                # Controlar Focos Derechos
                if estado_der == "NUEVO":
                    estado_gui["der_color"] = "#2ECC71" # Verde Neón
                    estado_gui["der_verde_ts"] = time.time()
                else:
                    estado_gui["der_color"] = "#E74C3C" # Rojo Neón

                # Guardar en Excel
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb.active
                hay_rechazo = (estado_der != "NUEVO") or (estado_izq != "NUEVO")
                
                registrar_en_excel(ws, resultado_der, "Derecho", estado_der, fecha_hora_actual)
                registrar_en_excel(ws, resultado_izq, "Izquierdo", estado_izq, fecha_hora_actual)
                wb.save(ARCHIVO_EXCEL)
                
                # Accionar PLC
                if not enviar_dato_d0(plc, "0200" if hay_rechazo else "0100", "Alarma/OK"): plc = None
                if plc: time.sleep(TIEMPO_ACTIVACION)
                if plc and not enviar_dato_d0(plc, "0000", "Limpiando D0"): plc = None
                    
            except PermissionError:
                estado_gui["mensaje_alerta"] = "¡ERROR!\nCierra el archivo Excel\npara continuar."
                time.sleep(2)
            except Exception as e:
                print(f"[ERROR EN LÓGICA] {e}")
                
            if plc:
                codigo_der = None
                codigo_izq = None

# ==========================================
# 6. INTERFAZ GRÁFICA (GUI) - MODO OSCURO
# ==========================================
def iniciar_interfaz():
    # Paleta de Colores Dark Mode
    BG_MAIN = "#121212"      # Fondo principal casi negro
    BG_PANEL = "#1E1E1E"     # Paneles de escáneres
    BG_BAR = "#000000"       # Barra superior negra pura
    FG_TEXT = "#E0E0E0"      # Texto blanco humo
    COLOR_APAGADO = "#444444" # Foco apagado (gris oscuro)

    root = tk.Tk()
    root.title("Panel de Control - Trazabilidad")
    root.geometry("900x600")
    root.configure(bg=BG_MAIN)

    fuente_titulo = font.Font(family="Helvetica", size=18, weight="bold")
    fuente_codigo = font.Font(family="Consolas", size=16, weight="bold")
    fuente_status = font.Font(family="Helvetica", size=12, weight="bold")

    # --- BARRA SUPERIOR (ESTADO) ---
    frame_status = tk.Frame(root, bg=BG_BAR, pady=10)
    frame_status.pack(fill=tk.X, side=tk.TOP)

    lbl_plc = tk.Label(frame_status, text="PLC: BUSCANDO...", bg=BG_BAR, fg="yellow", font=fuente_status)
    lbl_plc.pack(side=tk.LEFT, expand=True)

    lbl_izq = tk.Label(frame_status, text="ESCÁNER IZQ: BUSCANDO...", bg=BG_BAR, fg="yellow", font=fuente_status)
    lbl_izq.pack(side=tk.LEFT, expand=True)

    lbl_der = tk.Label(frame_status, text="ESCÁNER DER: BUSCANDO...", bg=BG_BAR, fg="yellow", font=fuente_status)
    lbl_der.pack(side=tk.LEFT, expand=True)

    # --- PANEL CENTRAL (LECTORES) ---
    frame_central = tk.Frame(root, bg=BG_MAIN)
    frame_central.pack(fill=tk.BOTH, expand=True, pady=20)

    # Lector Izquierdo
    frame_izq = tk.Frame(frame_central, bg=BG_PANEL, bd=2, relief=tk.RIDGE)
    frame_izq.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=10)
    tk.Label(frame_izq, text="ESCÁNER IZQUIERDO", bg=BG_PANEL, fg=FG_TEXT, font=fuente_titulo).pack(pady=10)
    canvas_izq = tk.Canvas(frame_izq, width=150, height=150, bg=BG_PANEL, highlightthickness=0)
    canvas_izq.pack(pady=10)
    led_izq = canvas_izq.create_oval(25, 25, 125, 125, fill=COLOR_APAGADO, outline="#000000", width=4)
    lbl_cod_izq = tk.Label(frame_izq, text="ESPERANDO PIEZA...", bg=BG_PANEL, font=fuente_codigo, fg=FG_TEXT, justify=tk.CENTER)
    lbl_cod_izq.pack(pady=20)

    # Lector Derecho
    frame_der = tk.Frame(frame_central, bg=BG_PANEL, bd=2, relief=tk.RIDGE)
    frame_der.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=10)
    tk.Label(frame_der, text="ESCÁNER DERECHO", bg=BG_PANEL, fg=FG_TEXT, font=fuente_titulo).pack(pady=10)
    canvas_der = tk.Canvas(frame_der, width=150, height=150, bg=BG_PANEL, highlightthickness=0)
    canvas_der.pack(pady=10)
    led_der = canvas_der.create_oval(25, 25, 125, 125, fill=COLOR_APAGADO, outline="#000000", width=4)
    lbl_cod_der = tk.Label(frame_der, text="ESPERANDO PIEZA...", bg=BG_PANEL, font=fuente_codigo, fg=FG_TEXT, justify=tk.CENTER)
    lbl_cod_der.pack(pady=20)

    # --- PANEL INFERIOR (KPIs) ---
    frame_kpi = tk.Frame(root, bg=BG_BAR, pady=15)
    frame_kpi.pack(fill=tk.X, side=tk.BOTTOM)
    
    lbl_kpi_nuevos = tk.Label(frame_kpi, text="NUEVOS: 0", bg=BG_BAR, fg="#2ECC71", font=fuente_titulo)
    lbl_kpi_nuevos.pack(side=tk.LEFT, expand=True)
    lbl_kpi_dup = tk.Label(frame_kpi, text="DUPLICADOS: 0", bg=BG_BAR, fg="#E74C3C", font=fuente_titulo)
    lbl_kpi_dup.pack(side=tk.LEFT, expand=True)
    lbl_kpi_err = tk.Label(frame_kpi, text="ERRORES: 0", bg=BG_BAR, fg="#F1C40F", font=fuente_titulo)
    lbl_kpi_err.pack(side=tk.LEFT, expand=True)

    # --- PANTALLA GIGANTE DE ALERTA (OCULTA POR DEFECTO) ---
    frame_alerta = tk.Frame(root, bg="#C0392B")
    lbl_alerta = tk.Label(frame_alerta, text="", bg="#C0392B", fg="white", font=("Arial", 28, "bold"), justify=tk.CENTER)
    lbl_alerta.pack(expand=True)

    # --- RUTINA DE ACTUALIZACIÓN (10 VECES POR SEGUNDO) ---
    def refrescar_pantalla():
        # Actualizar Conexiones
        lbl_plc.config(text="PLC: EN LÍNEA" if estado_gui["plc_ok"] else "PLC: DESCONECTADO", fg="#2ECC71" if estado_gui["plc_ok"] else "#E74C3C")
        lbl_izq.config(text="IZQ: EN LÍNEA" if estado_gui["izq_ok"] else "IZQ: DESCONECTADO", fg="#2ECC71" if estado_gui["izq_ok"] else "#E74C3C")
        lbl_der.config(text="DER: EN LÍNEA" if estado_gui["der_ok"] else "DER: DESCONECTADO", fg="#2ECC71" if estado_gui["der_ok"] else "#E74C3C")

        # Controlar Pantalla de Alerta Gigante
        if estado_gui["mensaje_alerta"]:
            lbl_alerta.config(text=estado_gui["mensaje_alerta"])
            frame_alerta.place(relx=0, rely=0, relwidth=1, relheight=1) # Cubre toda la pantalla
        else:
            frame_alerta.place_forget() # Oculta la pantalla

        # Apagar Foco Verde después de 3 segundos
        if estado_gui["izq_color"] == "#2ECC71" and (time.time() - estado_gui["izq_verde_ts"]) >= 3.0:
            estado_gui["izq_color"] = COLOR_APAGADO
            estado_gui["izq_codigo"] = "ESPERANDO PIEZA..."
            
        if estado_gui["der_color"] == "#2ECC71" and (time.time() - estado_gui["der_verde_ts"]) >= 3.0:
            estado_gui["der_color"] = COLOR_APAGADO
            estado_gui["der_codigo"] = "ESPERANDO PIEZA..."

        # Actualizar Focos y Textos (Si es rojo se queda así, el verde desaparece)
        canvas_izq.itemconfig(led_izq, fill=estado_gui["izq_color"])
        lbl_cod_izq.config(text=estado_gui["izq_codigo"], fg="#E74C3C" if estado_gui["izq_color"] == "#E74C3C" else FG_TEXT)

        canvas_der.itemconfig(led_der, fill=estado_gui["der_color"])
        lbl_cod_der.config(text=estado_gui["der_codigo"], fg="#E74C3C" if estado_gui["der_color"] == "#E74C3C" else FG_TEXT)

        # Actualizar Contadores
        lbl_kpi_nuevos.config(text=f"NUEVOS: {estado_gui['kpi_nuevos']}")
        lbl_kpi_dup.config(text=f"DUPLICADOS: {estado_gui['kpi_dup']}")
        lbl_kpi_err.config(text=f"ERRORES: {estado_gui['kpi_err']}")

        root.after(100, refrescar_pantalla) # Repite la función cada 100 milisegundos

    refrescar_pantalla()
    root.mainloop()

# ==========================================
# 7. ARRANQUE DEL SISTEMA
# ==========================================
if __name__ == "__main__":
    # Inicia la lógica de escaneo en un hilo invisible (Background)
    hilo = threading.Thread(target=hilo_produccion, daemon=True)
    hilo.start()
    
    # Inicia la ventana gráfica
    iniciar_interfaz()